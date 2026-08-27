# -*- coding: utf-8 -*-
"""Build the monthly .xlsx report.

A human-reviewable companion to the two raw upload CSVs: a Summary sheet
matching the email report, the QA Checks results colour-coded pass/fail, a
generated Label Taxonomy reference, and both feeds as real, filterable
sheets. Committed to the repo (so every month's report is in git history)
and attached to the report email alongside the CSVs.

openpyxl is the one deliberate exception to this project's
no-third-party-dependencies rule — see DECISIONS.md D13. It only ever
writes a local file and never touches the network, so it carries none of
the risk D3/D7 documented for the fetching layer.
"""
import os

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

import config as cfg

HEADER_FILL = PatternFill(start_color="0C6776", end_color="0C6776", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
PASS_FILL = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
FAIL_FILL = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")


def _header_row(ws, headers):
    ws.append(list(headers))
    for cell in ws[ws.max_row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    ws.freeze_panes = "A2"


def _autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _section(ws, title, rows):
    ws.append([title])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=12)
    for row in rows:
        ws.append(list(row))
    ws.append([])


def _summary_sheet(wb, summary):
    ws = wb.active
    ws.title = "Summary"
    ws.append(["Riparide page feed refresh"])
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws.append(["Run", summary.get("run_at", "")])
    ws.append([])

    if summary.get("failed"):
        _section(ws, "RUN FAILED - no files were produced", [("Reason", summary["failed"])])
        for row in ws.iter_rows(min_row=ws.max_row - 2, max_row=ws.max_row - 2):
            for cell in row:
                cell.fill = FAIL_FILL

    d = summary.get("diff") or {}
    if d:
        _section(ws, "Changes", [
            ("Added", len(d.get("added", []))),
            ("Removed", len(d.get("removed", []))),
            ("Unchanged", d.get("unchanged_count", 0)),
            ("Feed size", "%d -> %d" % (d.get("previous_total", 0), d.get("current_total", 0))),
        ])

    st = summary.get("status") or {}
    if st:
        rows = [("Checked", st.get("checked", 0)), ("Excluded (non-200)", st.get("excluded", 0))]
        if st.get("unchecked"):
            rows.append(("Unchecked (above cap)", st["unchecked"]))
        _section(ws, "Status checks", rows)

    loc = summary.get("location") or {}
    if loc and not loc.get("skipped"):
        _section(ws, "Location", [
            ("Filled this run", loc.get("applied", 0)),
            ("Pages read this run", loc.get("fetched", 0)),
            ("Cached in total", loc.get("cached_total", 0)),
            ("Still without location", loc.get("still_without_location", 0)),
        ])

    rb = summary.get("robots") or {}
    if rb:
        if not rb.get("fetched"):
            rows = [("Could not check", rb.get("note", ""))]
        elif rb.get("blocked"):
            rows = [("WARNING: AdsBot blocked",
                     "; ".join("%s by Disallow: %s" % (p, dd) for p, dd in rb["blocked"]))]
        else:
            rows = [("AdsBot-Google / AdsBot-Google-Mobile", "not blocked")]
        _section(ws, "robots.txt / AdsBot check", rows)

    out = summary.get("outputs") or {}
    if out:
        _section(ws, "Files", [(feed, n, os.path.basename(path)) for feed, (path, n) in out.items()])

    _autosize(ws, [32, 45, 30])


def _qa_sheet(wb, checks):
    ws = wb.create_sheet("QA Checks")
    _header_row(ws, ["Check", "Result", "Detail", "Threshold"])
    for c in checks:
        ws.append([c["check"], "PASS" if c["passed"] else "FAIL", c["detail"], c["threshold"]])
        fill = PASS_FILL if c["passed"] else FAIL_FILL
        for cell in ws[ws.max_row]:
            cell.fill = fill
    _autosize(ws, [58, 10, 55, 30])


def _taxonomy_sheet(wb):
    """A living reference generated from config.py, not hand-maintained."""
    ws = wb.create_sheet("Label Taxonomy")
    _header_row(ws, ["Dimension", "Label", "Notes"])

    rows = [("Page type", pt, "") for pt in (
        "PAGE_LISTING", "PAGE_STORY", "PAGE_REGION", "PAGE_STATE", "PAGE_COUNTRY",
        "PAGE_FACET_TYPE", "PAGE_ADVENTURE", "PAGE_COLLECTION", "PAGE_COLLECTION_HUB",
        "PAGE_CORE (excluded from both feeds)")]

    seen_geo = set()
    for country, state, geo_c, geo_s in cfg.FACET_SCOPES:
        if geo_c not in seen_geo:
            seen_geo.add(geo_c)
            rows.append(("Geo (country)", geo_c, country))
        if geo_s and geo_s not in seen_geo:
            seen_geo.add(geo_s)
            rows.append(("Geo (state)", geo_s, "%s, %s" % (country, state)))

    for sub in cfg.SUBCATS:
        rows.append(("Stay type", "TYPE_" + sub.upper().replace("-", "_"), "facet subcategory: " + sub))
    for wording, label in cfg.EXTRA_STAY_WORDINGS.items():
        rows.append(("Stay type (no facet page)", label, "title wording: \"%s\"" % wording))

    seen_intent = set()
    for kw, label in cfg.INTENT_KEYWORDS:
        if label in seen_intent:
            continue
        seen_intent.add(label)
        keywords = ", ".join(k for k, l in cfg.INTENT_KEYWORDS if l == label)
        head = " (SEARCH_HEAD boundary)" if label in cfg.SEARCH_HEAD_INTENTS else ""
        rows.append(("Intent", label + head, "keyword(s): " + keywords))

    rows.extend([
        ("Boundary", "SEARCH_HEAD", "Search Generic beats PMax 2x+ here; excluded from PMax under Option B"),
        ("Boundary", "PMAX_LONGTAIL", "PMax's proven territory"),
        ("Boundary", "NZ_REVIEW", "NZ PMax at 0.68 ROAS; held out of asset groups"),
        ("Boundary", "US_HOLD", "US is not a paid market; held out of asset groups"),
        ("Boundary", "ADV_HOLD", "Editorial adventure pages; held out of asset groups"),
    ])

    for row in rows:
        ws.append(list(row))
    _autosize(ws, [24, 32, 60])


def _feed_sheet(wb, title, rows):
    ws = wb.create_sheet(title)
    _header_row(ws, cfg.CSV_HEADER)
    for r in rows:
        ws.append([r["url"], r["label"]])
    _autosize(ws, [75, 65])


def write_report(out_rows, checks, summary, path=None):
    """Build the workbook and write it atomically. Returns the path written."""
    path = path or os.path.join(cfg.REPORT_DIR, cfg.REPORT_XLSX)
    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)

    wb = Workbook()
    _summary_sheet(wb, summary)
    _qa_sheet(wb, checks)
    _taxonomy_sheet(wb)
    _feed_sheet(wb, "Page Feed - Core", [r for r in out_rows if r["feed"] == cfg.FEED_CORE])
    _feed_sheet(wb, "Page Feed - Adventures", [r for r in out_rows if r["feed"] == cfg.FEED_ADVENTURES])

    tmp = path + ".tmp"
    wb.save(tmp)
    os.replace(tmp, path)
    return path
