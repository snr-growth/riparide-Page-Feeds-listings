# -*- coding: utf-8 -*-
"""Unit tests for report.py's .xlsx generation."""
import os
import shutil
import tempfile
import unittest

from openpyxl import load_workbook

import config as cfg
import report

SAMPLE_SUMMARY = {
    "run_at": "2026-09-01 03:00 UTC",
    "failed": None,
    "diff": {"added": ["https://x/1"], "removed": [], "unchanged_count": 5,
             "previous_total": 5, "current_total": 6},
    "status": {"checked": 1, "excluded": 0},
    "location": {"applied": 1, "fetched": 1, "cached_total": 10, "still_without_location": 0},
    "robots": {"fetched": True, "adsbot_group_found": False, "blocked": [],
               "note": "no AdsBot-specific group found"},
    "outputs": {cfg.FEED_CORE: ("data/output/core.csv", 2), cfg.FEED_ADVENTURES: ("data/output/adv.csv", 1)},
}

SAMPLE_CHECKS = [
    {"check": "Duplicate URLs across both feeds", "passed": True, "detail": "0 duplicated", "threshold": "0"},
    {"check": "Feed size has not collapsed versus last run", "passed": False,
     "detail": "CORE: 100 -> 3", "threshold": ">= 50% of last run's row count"},
]

SAMPLE_ROWS = [
    {"url": "https://www.riparide.com/listings/1-x", "label": "PAGE_LISTING;PMAX_LONGTAIL", "feed": cfg.FEED_CORE},
    {"url": "https://www.riparide.com/adventures/2-y", "label": "PAGE_ADVENTURE;ADV_HOLD",
     "feed": cfg.FEED_ADVENTURES},
]


class WriteReportTests(unittest.TestCase):
    def setUp(self):
        self.dir = tempfile.mkdtemp()
        self.path = os.path.join(self.dir, "report.xlsx")

    def tearDown(self):
        shutil.rmtree(self.dir, ignore_errors=True)

    def test_writes_a_readable_file_with_expected_sheets(self):
        out = report.write_report(SAMPLE_ROWS, SAMPLE_CHECKS, SAMPLE_SUMMARY, path=self.path)
        self.assertEqual(out, self.path)
        wb = load_workbook(self.path)
        self.assertEqual(wb.sheetnames,
                         ["Summary", "QA Checks", "Label Taxonomy",
                          "Page Feed - Core", "Page Feed - Adventures"])

    def test_summary_sheet_reflects_the_run(self):
        report.write_report(SAMPLE_ROWS, SAMPLE_CHECKS, SAMPLE_SUMMARY, path=self.path)
        wb = load_workbook(self.path)
        values = [c[0] for c in wb["Summary"].iter_rows(values_only=True) if c[0]]
        self.assertIn("Changes", values)
        self.assertIn("Added", values)

    def test_qa_sheet_has_header_and_one_row_per_check(self):
        report.write_report(SAMPLE_ROWS, SAMPLE_CHECKS, SAMPLE_SUMMARY, path=self.path)
        ws = load_workbook(self.path)["QA Checks"]
        self.assertEqual(ws.max_row, 1 + len(SAMPLE_CHECKS))
        header = [c.value for c in ws[1]]
        self.assertEqual(header, ["Check", "Result", "Detail", "Threshold"])
        results = [row[1].value for row in ws.iter_rows(min_row=2)]
        self.assertEqual(results, ["PASS", "FAIL"])

    def test_feed_sheets_contain_only_their_own_rows(self):
        report.write_report(SAMPLE_ROWS, SAMPLE_CHECKS, SAMPLE_SUMMARY, path=self.path)
        wb = load_workbook(self.path)
        core_urls = [r[0] for r in wb["Page Feed - Core"].iter_rows(min_row=2, values_only=True)]
        adv_urls = [r[0] for r in wb["Page Feed - Adventures"].iter_rows(min_row=2, values_only=True)]
        self.assertEqual(core_urls, ["https://www.riparide.com/listings/1-x"])
        self.assertEqual(adv_urls, ["https://www.riparide.com/adventures/2-y"])

    def test_taxonomy_sheet_includes_getaway_and_is_generated_not_hand_written(self):
        report.write_report(SAMPLE_ROWS, SAMPLE_CHECKS, SAMPLE_SUMMARY, path=self.path)
        ws = load_workbook(self.path)["Label Taxonomy"]
        labels = [row[1] for row in ws.iter_rows(min_row=2, values_only=True)]
        self.assertIn("INT_GETAWAY (SEARCH_HEAD boundary)", labels)
        self.assertIn("SEARCH_HEAD", labels)
        # every SUBCATS entry must be represented so the sheet can't silently
        # drift out of sync with config.py
        for sub in cfg.SUBCATS:
            self.assertIn("TYPE_" + sub.upper().replace("-", "_"), labels)

    def test_failed_run_summary_is_still_readable(self):
        summary = dict(SAMPLE_SUMMARY, failed="RuntimeError: validation failed, no files were written")
        report.write_report([], [], summary, path=self.path)
        wb = load_workbook(self.path)
        values = [c[0] for c in wb["Summary"].iter_rows(values_only=True) if c[0]]
        self.assertTrue(any("RUN FAILED" in str(v) for v in values))

    def test_empty_feeds_produce_header_only_sheets(self):
        report.write_report([], SAMPLE_CHECKS, SAMPLE_SUMMARY, path=self.path)
        wb = load_workbook(self.path)
        self.assertEqual(wb["Page Feed - Core"].max_row, 1)
        self.assertEqual(wb["Page Feed - Adventures"].max_row, 1)

    def test_creates_parent_directory(self):
        nested = os.path.join(self.dir, "nested", "dir", "report.xlsx")
        report.write_report(SAMPLE_ROWS, SAMPLE_CHECKS, SAMPLE_SUMMARY, path=nested)
        self.assertTrue(os.path.exists(nested))


if __name__ == "__main__":
    unittest.main()
